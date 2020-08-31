# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'reversion2.ui'
#
# Created by: PyQt5 UI code generator 5.11.2
#
# WARNING! All changes made in this file will be lost!

#PyQT Designer에서 기본적으로 import하는 것들
from PyQt5 import QtCore, QtGui, QtWidgets
#메뉴바 혹은 상태표시줄을 만들기 위해서는 QMainWindow를 사용한다.
#QAction은 메뉴바에서 특정 항목을 선택하여 그 내부 그룹의 내용들이 어떤 행동을 하도록 만든다.
from PyQt5.QtWidgets import QApplication, QMainWindow, QAction, QMenu, QTextEdit, QFileDialog, QTextBrowser
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QIcon
from PyQt5.Qt import QColor
from PyQt5.QtCore import *
import numpy as np
import sys, random

import cm1k_emulator as cm1k
import csv
import openpyxl
import test
#import log
import neuron as nrn
import pyqtgraph as pg
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as Canvas
import matplotlib
from PIL import ImageGrab
import time
from pynput.mouse import Button, Controller

#Ensure using PyQt5 backend
matplotlib.use('QT5Agg')

#######에러 찾기############
sys._excepthook = sys.excepthook
def my_exception_hook(exctype, value, traceback):
    # Print the error and traceback
    print(exctype, value, traceback)
    # Call the normal Exception hook after
    sys._excepthook(exctype, value, traceback)
    sys.exit(1)
# Set the exception hook to our wrapping function
sys.excepthook = my_exception_hook
try:
    sys.exit(app.exec_())
except:
    print("Exiting")
##############################

#Ui를 정의하고 있는 클래스
class Ui_MainWindow(QtWidgets.QMainWindow):
    def setupUi(self, MainWindow):
        self.load_data  = []
        self.neurons = []
        self.log = []
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1435,1036)
        #central widget
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        #graph
        self.mainGraph = MplWidget(self.centralwidget)
        self.mainGraph.setGeometry(QtCore.QRect(710, 50, 711, 311))
        self.mainGraph.setObjectName("mainGraph")
        self.neuronsGraph = MplWidget(self.centralwidget)
        self.neuronsGraph.setGeometry(QtCore.QRect(710, 370, 351, 301))
        self.neuronsGraph.setObjectName("neuronsGraph")
        self.learningCurveGraph = MplWidget(self.centralwidget)
        self.learningCurveGraph.setGeometry(QtCore.QRect(1070, 370, 351, 301))
        self.learningCurveGraph.setObjectName("learningCurveGraph")
        self.accuracyCategoryBarGraph = MplWidget(self.centralwidget)
        self.accuracyCategoryBarGraph.setGeometry(QtCore.QRect(710, 680, 351, 301))
        self.accuracyCategoryBarGraph.setObjectName("accuracyCategoryBarGraph")
        self.accuracyPieGraph=MplWidget(self.centralwidget)
        self.accuracyPieGraph.setGeometry(QtCore.QRect(1070, 680, 351, 301))
        self.accuracyPieGraph.setObjectName("accuracyPieGraph")
        #group box
        self.StatusBox = QtWidgets.QGroupBox(self.centralwidget)
        self.StatusBox.setGeometry(QtCore.QRect(10, 30, 331, 331))
        self.StatusBox.setObjectName("StatusBox")
        self.ButtonBox = QtWidgets.QGroupBox(self.centralwidget)
        self.ButtonBox.setGeometry(QtCore.QRect(350, 30, 331, 331))
        self.ButtonBox.setTitle("")
        self.ButtonBox.setObjectName("ButtonBox")
        self.learningGroupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.learningGroupBox.setGeometry(QtCore.QRect(10, 370, 331, 271))
        self.learningGroupBox.setObjectName("learningGroupBox")
        self.classificationGroupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.classificationGroupBox.setGeometry(QtCore.QRect(350, 370, 331, 271))
        self.classificationGroupBox.setObjectName("classificationGroupBox")
        self.ResultBox = QtWidgets.QGroupBox(self.centralwidget)
        self.ResultBox.setGeometry(QtCore.QRect(350, 650, 331, 331))
        self.ResultBox.setObjectName("ResultBox")
        #push button
        self.pushButton_LoadDataset = QtWidgets.QPushButton(self.ButtonBox)
        self.pushButton_LoadDataset.setGeometry(QtCore.QRect(10, 10, 91, 91))
        # self.pushButton.setIconSize(QtCore.QSize(31, 30))
        self.pushButton_LoadDataset.setObjectName("LoadDataset")
        self.pushButton_ExportDataset = QtWidgets.QPushButton(self.ButtonBox)
        self.pushButton_ExportDataset.setGeometry(QtCore.QRect(110, 10, 91, 91))
        self.pushButton_ExportDataset.setObjectName("ExportDataset")
        self.pushButton_LoadKnowledge = QtWidgets.QPushButton(self.ButtonBox)
        self.pushButton_LoadKnowledge.setGeometry(QtCore.QRect(210, 10, 91, 91))
        self.pushButton_LoadKnowledge.setObjectName("LoadKnowledge")
        self.pushButton_SaveKnowledge = QtWidgets.QPushButton(self.ButtonBox)
        self.pushButton_SaveKnowledge.setGeometry(QtCore.QRect(10, 110, 91, 91))
        self.pushButton_SaveKnowledge.setCursor(QtGui.QCursor(QtCore.Qt.WaitCursor))
        self.pushButton_SaveKnowledge.setObjectName("SaveKnowledge")
        self.pushButton_Knowledge = QtWidgets.QPushButton(self.ButtonBox)
        self.pushButton_Knowledge.setGeometry(QtCore.QRect(110, 110, 91, 91))
        self.pushButton_Knowledge.setObjectName("Knowledge")
        self.pushButton_DatasetInDetail = QtWidgets.QPushButton(self.ButtonBox)
        self.pushButton_DatasetInDetail.setGeometry(QtCore.QRect(210, 110, 91, 91))
        self.pushButton_DatasetInDetail.setObjectName("DatasetInDetail")
        self.pushButton_ForgetAll = QtWidgets.QPushButton(self.ButtonBox)
        self.pushButton_ForgetAll.setGeometry(QtCore.QRect(110, 210, 91, 91))
        self.pushButton_ForgetAll.setObjectName("ForgetAll")
        self.pushButton_Learn = QtWidgets.QPushButton(self.learningGroupBox)
        self.pushButton_Learn.setGeometry(QtCore.QRect(220, 160, 91, 91))
        self.pushButton_Learn.setObjectName("pushButton_Learn")
        self.pushButton_Classify = QtWidgets.QPushButton(self.classificationGroupBox)
        self.pushButton_Classify.setGeometry(QtCore.QRect(220, 160, 91, 91))
        self.pushButton_Classify.setObjectName("pushButton_Classify")
        self.pushButton_Exit = QtWidgets.QPushButton(self.ResultBox)
        self.pushButton_Exit.setGeometry(QtCore.QRect(220, 220, 91, 91))
        self.pushButton_Exit.setObjectName("pushButton_Exit")
        self.pushButton_repeat = QtWidgets.QPushButton(self.ButtonBox)
        self.pushButton_repeat.setGeometry(QtCore.QRect(10, 210, 91, 91))
        self.pushButton_repeat.setObjectName("pushButton_Repeat")
        #radio button
        self.radioButton_Iterative = QtWidgets.QRadioButton(self.learningGroupBox)
        self.radioButton_Iterative.setGeometry(QtCore.QRect(10, 15, 161, 21))
        self.radioButton_Iterative.setObjectName("radioButton_Iterative")
        self.radioButton_Iterative.setProperty("checked", True)
        # self.radioButton_DeepRBF = QtWidgets.QRadioButton(self.learningGroupBox)
        # self.radioButton_DeepRBF.setGeometry(QtCore.QRect(10, 60, 90, 16))
        # self.radioButton_DeepRBF.setObjectName("radioButton_DeepRBF")
        self.radioButton_WriteAllSamples = QtWidgets.QRadioButton(self.learningGroupBox)
        self.radioButton_WriteAllSamples.setGeometry(QtCore.QRect(10, 80, 141, 16))
        self.radioButton_WriteAllSamples.setObjectName("radioButton_WriteAllSamples")
        self.radioButton_Classify1=QtWidgets.QRadioButton(self.classificationGroupBox)
        self.radioButton_Classify1.setGeometry(QtCore.QRect(10, 20, 220, 16))
        self.radioButton_Classify1.setObjectName("radioButton_Classify1")
        self.radioButton_Classify1.setProperty("checked", True)
        self.radioButton_Classify2=QtWidgets.QRadioButton(self.classificationGroupBox)
        self.radioButton_Classify2.setGeometry(QtCore.QRect(10, 40, 220, 16))
        self.radioButton_Classify2.setObjectName("radioButton_Classify2")
        self.radioButton_Classify3 = QtWidgets.QRadioButton(self.classificationGroupBox)
        self.radioButton_Classify3.setGeometry(QtCore.QRect(10, 60, 220, 16))
        self.radioButton_Classify3.setObjectName("radioButton_Classify3")
        self.radioButton_Classify4 = QtWidgets.QRadioButton(self.classificationGroupBox)
        self.radioButton_Classify4.setGeometry(QtCore.QRect(10, 80, 220, 16))
        self.radioButton_Classify4.setObjectName("radioButton_Classify4")
        self.radioButton_Classify5 = QtWidgets.QRadioButton(self.classificationGroupBox)
        self.radioButton_Classify5.setGeometry(QtCore.QRect(10, 100, 220, 16))
        self.radioButton_Classify5.setObjectName("radioButton_Classify5")
        self.radioButton_Classify6 = QtWidgets.QRadioButton(self.classificationGroupBox)
        self.radioButton_Classify6.setGeometry(QtCore.QRect(10, 120, 220, 16))
        self.radioButton_Classify6.setObjectName("radioButton_Classify6")
        self.radioButton_Classify7 = QtWidgets.QRadioButton(self.classificationGroupBox)
        self.radioButton_Classify7.setGeometry(QtCore.QRect(10, 140, 220, 16))
        self.radioButton_Classify7.setObjectName("radioButton_Classify7")
        self.radioButton_Classify8 = QtWidgets.QRadioButton(self.classificationGroupBox)
        self.radioButton_Classify8.setGeometry(QtCore.QRect(10, 160, 220, 16))
        self.radioButton_Classify8.setObjectName("radioButton_Classify8")
        #spin box
        self.spinBox_SimuNMnK = QtWidgets.QSpinBox(self.StatusBox)
        self.spinBox_SimuNMnK.setGeometry(QtCore.QRect(150, 200, 42, 22))
        self.spinBox_SimuNMnK.setProperty("value", 1)
        self.spinBox_SimuNMnK.setObjectName("spinBox_SimuNMnK")
        self.spinBox_Iter = QtWidgets.QSpinBox(self.learningGroupBox)
        self.spinBox_Iter.setGeometry(QtCore.QRect(110, 40, 42, 16))
        self.spinBox_Iter.setObjectName("spinBox_Iter")
        self.spinBox_maxif = QtWidgets.QSpinBox(self.learningGroupBox)
        self.spinBox_maxif.setGeometry(QtCore.QRect(80, 120, 81, 22))
        self.spinBox_maxif.setMaximum(16384)
        self.spinBox_maxif.setMinimum(2)
        self.spinBox_maxif.setProperty("value", 16384)
        self.spinBox_maxif.setObjectName("spinBox_maxif")
        self.spinBox_minif = QtWidgets.QSpinBox(self.learningGroupBox)
        self.spinBox_minif.setGeometry(QtCore.QRect(80, 150, 81, 22))
        self.spinBox_minif.setMaximum(16384)
        self.spinBox_minif.setMinimum(2)
        self.spinBox_minif.setProperty("value", 2)
        self.spinBox_minif.setObjectName("spinBox_minif")
        self.spinBox_Kvalue = QtWidgets.QSpinBox(self.classificationGroupBox)
        self.spinBox_Kvalue.setGeometry(QtCore.QRect(40, 180, 42, 22))
        self.spinBox_Kvalue.setProperty("value", 3)
        self.spinBox_Kvalue.setObjectName("spinBox_Kvalue")
        self.spinBox_MinConsensus = QtWidgets.QSpinBox(self.classificationGroupBox)
        self.spinBox_MinConsensus.setGeometry(QtCore.QRect(175, 80, 30, 16))
        self.spinBox_MinConsensus.setProperty("value", 2)
        self.spinBox_MinConsensus.setObjectName("spinBox_MinConsensus")
        self.spinBox_MinConsensus2 = QtWidgets.QSpinBox(self.classificationGroupBox)
        self.spinBox_MinConsensus2.setGeometry(QtCore.QRect(175, 160, 30, 16))
        self.spinBox_MinConsensus2.setProperty("value", 2)
        self.spinBox_MinConsensus2.setObjectName("spinBox_MinConsensus2")
        #log text browser
        self.logBrowser = QtWidgets.QTextBrowser(self.centralwidget)
        self.logBrowser.setGeometry(QtCore.QRect(10, 670, 331, 311))
        self.logBrowser.setObjectName("logBrowser")
        self.Browser_DName = QtWidgets.QTextBrowser(self.StatusBox)
        self.Browser_DName.setGeometry(QtCore.QRect(150, 30, 104, 31))
        self.Browser_DName.setObjectName("Browser_DName")
        self.Browser_NSamples = QtWidgets.QTextBrowser(self.StatusBox)
        self.Browser_NSamples.setGeometry(QtCore.QRect(150, 60, 104, 31))
        self.Browser_NSamples.setObjectName("Browser_NSamples")
        self.Browser_NNeurons = QtWidgets.QTextBrowser(self.StatusBox)
        self.Browser_NNeurons.setGeometry(QtCore.QRect(150, 90, 104, 31))
        self.Browser_NNeurons.setObjectName("Browser_NNeurons")
        self.Browser_NNeurons.append("1024")
        self.Browser_UsedNeurons = QtWidgets.QTextBrowser(self.StatusBox)
        self.Browser_UsedNeurons.setGeometry(QtCore.QRect(150, 120, 104, 31))
        self.Browser_UsedNeurons.setObjectName("Browser_UsedNeurons")
        self.Browser_NSamples2 = QtWidgets.QTextBrowser(self.ResultBox)
        self.Browser_NSamples2.setGeometry(QtCore.QRect(140, 30, 104, 31))
        self.Browser_NSamples2.setObjectName("Browser_NSamples2")
        self.Browser_ID = QtWidgets.QTextBrowser(self.ResultBox)
        self.Browser_ID.setGeometry(QtCore.QRect(140, 60, 104, 31))
        self.Browser_ID.setObjectName("Browser_ID")
        self.Browser_UNCc = QtWidgets.QTextBrowser(self.ResultBox)
        self.Browser_UNCc.setGeometry(QtCore.QRect(140, 90, 104, 31))
        self.Browser_UNCc.setObjectName("Browser_UNCc")
        self.Browser_UNCi = QtWidgets.QTextBrowser(self.ResultBox)
        self.Browser_UNCi.setGeometry(QtCore.QRect(140, 120, 104, 31))
        self.Browser_UNCi.setObjectName("Browser_UNCi")
        self.Browser_UNK = QtWidgets.QTextBrowser(self.ResultBox)
        self.Browser_UNK.setGeometry(QtCore.QRect(140, 150, 104, 31))
        self.Browser_UNK.setObjectName("Browser_UNK")
        self.Browser_Total = QtWidgets.QTextBrowser(self.ResultBox)
        self.Browser_Total.setGeometry(QtCore.QRect(140, 180, 104, 31))
        self.Browser_Total.setObjectName("Browser_Total")
        self.Browser_Count = QtWidgets.QTextBrowser(self.StatusBox)
        self.Browser_Count.setGeometry(QtCore.QRect(120, 260, 41, 31))
        self.Browser_Count.setObjectName("Browser_Count")
        #check box - Max_Iter.
        self.checkBox = QtWidgets.QCheckBox(self.learningGroupBox)
        self.checkBox.setGeometry(QtCore.QRect(30, 36, 81, 20))
        self.checkBox.setObjectName("checkBox")
        #font
        font = QtGui.QFont()
        font.setFamily("새굴림")
        font.setPointSize(15)
        #label
        self.label_log = QtWidgets.QLabel(self.centralwidget)
        self.label_log.setGeometry(QtCore.QRect(10, 650, 56, 12))
        self.label_log.setObjectName("label_log")
        self.label_MinIF = QtWidgets.QLabel(self.learningGroupBox)
        self.label_MinIF.setGeometry(QtCore.QRect(10, 150, 61, 16))
        self.label_MinIF.setObjectName("label_MinIF")
        self.label_MaxIF = QtWidgets.QLabel(self.learningGroupBox)
        self.label_MaxIF.setGeometry(QtCore.QRect(10, 122, 61, 20))
        self.label_MaxIF.setObjectName("label_MaxIF")
        self.label_K = QtWidgets.QLabel(self.classificationGroupBox)
        self.label_K.setGeometry(QtCore.QRect(20, 180, 21, 21))
        self.label_K.setObjectName("label_K")
        self.label_DName = QtWidgets.QLabel(self.StatusBox)
        self.label_DName.setGeometry(QtCore.QRect(30, 30, 121, 31))
        self.label_DName.setObjectName("label_DName")
        self.label_NSamples = QtWidgets.QLabel(self.StatusBox)
        self.label_NSamples.setGeometry(QtCore.QRect(30, 60, 101, 31))
        self.label_NSamples.setObjectName("label_NSamples")
        self.label_NNeurons = QtWidgets.QLabel(self.StatusBox)
        self.label_NNeurons.setGeometry(QtCore.QRect(30, 90, 81, 31))
        self.label_NNeurons.setObjectName("label_NNeurons")
        self.label_UsedNeurons = QtWidgets.QLabel(self.StatusBox)
        self.label_UsedNeurons.setGeometry(QtCore.QRect(30, 120, 121, 31))
        self.label_UsedNeurons.setObjectName("label_UsedNeurons")
        self.label_MnK = QtWidgets.QLabel(self.StatusBox)
        self.label_MnK.setGeometry(QtCore.QRect(30, 200, 81, 21))
        self.label_MnK.setObjectName("label_MnK")
        self.label_n = QtWidgets.QLabel(self.StatusBox)
        self.label_n.setGeometry(QtCore.QRect(120, 200, 21, 21))
        self.label_n.setObjectName("label_n")
        self.label_NSamples2 = QtWidgets.QLabel(self.ResultBox)
        self.label_NSamples2.setGeometry(QtCore.QRect(20, 30, 121, 31))
        self.label_NSamples2.setObjectName("label_NSamples2")
        self.label_ID = QtWidgets.QLabel(self.ResultBox)
        self.label_ID.setGeometry(QtCore.QRect(20, 60, 121, 31))
        self.label_ID.setObjectName("label_ID")
        self.label_UNCc = QtWidgets.QLabel(self.ResultBox)
        self.label_UNCc.setGeometry(QtCore.QRect(20, 90, 121, 31))
        self.label_UNCc.setObjectName("label_UNCc")
        self.label_UNCi = QtWidgets.QLabel(self.ResultBox)
        self.label_UNCi.setGeometry(QtCore.QRect(20, 120, 121, 31))
        self.label_UNCi.setObjectName("label_UNCi")
        self.label_UNK = QtWidgets.QLabel(self.ResultBox)
        self.label_UNK.setGeometry(QtCore.QRect(20, 150, 121, 31))
        self.label_UNK.setObjectName("label_UNK")
        self.label_Total = QtWidgets.QLabel(self.ResultBox)
        self.label_Total.setGeometry(QtCore.QRect(20, 180, 121, 31))
        self.label_Total.setObjectName("label_Total")
        self.label_MainGraph = QtWidgets.QLabel(self.centralwidget)
        self.label_MainGraph.setGeometry(QtCore.QRect(710,30,261,16))
        self.label_MainGraph.setObjectName("label_MainGraph")
        self.label_NeuronsGraph = QtWidgets.QLabel(self.centralwidget)
        self.label_NeuronsGraph.setGeometry(QtCore.QRect(720,370,261,21))
        self.label_NeuronsGraph.setObjectName("label_NeuronsGraph")
        self.label_LearningCurveGraph = QtWidgets.QLabel(self.centralwidget)
        self.label_LearningCurveGraph.setGeometry(QtCore.QRect(1080,370,261,21))
        self.label_LearningCurveGraph.setObjectName("label_LearningCurveGraph")
        self.label_AccuracyGraph = QtWidgets.QLabel(self.centralwidget)
        self.label_AccuracyGraph.setGeometry(QtCore.QRect(720,680,261,21))
        self.label_AccuracyGraph.setObjectName("label_AccuracyGraph")
        self.label_PieGraph = QtWidgets.QLabel(self.centralwidget)
        self.label_PieGraph.setGeometry(QtCore.QRect(1080,680,261,21))
        self.label_PieGraph.setObjectName("label_PieGraph")
        self.label_autoCounter = QtWidgets.QLabel(self.StatusBox)
        self.label_autoCounter.setGeometry(QtCore.QRect(70,270,161,21))
        self.label_autoCounter.setObjectName("label_autoCounter")
        MainWindow.setCentralWidget(self.centralwidget)
        #
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        #=========================================
        # 클릭 이벤트 생성
        #=========================================
        #==========pushButton================
        self.pushButton_ForgetAll.clicked.connect(self.ForgetAll_clicked)
        self.pushButton_Learn.clicked.connect(self.Learn_clicked)
        self.pushButton_Classify.clicked.connect(self.Classify_clicked)
        self.pushButton_LoadDataset.clicked.connect(self.showDialog)
        self.pushButton_ExportDataset.clicked.connect(self.showDialog2)
        self.pushButton_DatasetInDetail.clicked.connect(self.viewDetail)
        self.pushButton_Knowledge.clicked.connect(self.viewKnowledge)
        self.pushButton_SaveKnowledge.clicked.connect(self.SaveKnowledge)
        self.pushButton_LoadKnowledge.clicked.connect(self.LoadKnowledge)
        self.pushButton_Exit.clicked.connect(self.exitProgram)
        # self.pushButton_repeat.clicked.connect(self.repeat)
        #=========radioButton================
        self.radioButton_Iterative.clicked.connect(self.rdobtn_clicked)
        self.radioButton_WriteAllSamples.clicked.connect(self.rdobtn_clicked)
        # self.radioButton_DeepRBF.clicked.connect(self.rdobtn_clicked)
        #==========spinBox===================
        self.spinBox_maxif.valueChanged.connect(self.spinBoxChanged)
        self.spinBox_SimuNMnK.valueChanged.connect(self.spinBoxChanged)
        self.spinBox_Iter.valueChanged.connect(self.spinBoxChanged)
        self.spinBox_Kvalue.valueChanged.connect(self.spin_value)
        # checkBox
        #self.checkBox.stateChanged.connect(self.rdobtn_clicked)
        self.count = 0
        self.i=2
        self.Browser_Count.setText('0')

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Pre-Simulator"))
        self.label_log.setText(_translate("MainWindow", "Log"))
        self.learningGroupBox.setTitle(_translate("MainWindow", "RBF Learning Option"))
        self.label_MinIF.setText(_translate("MainWindow", "MinIF"))
        self.label_MaxIF.setText(_translate("MainWindow", "MaxIF"))
        self.radioButton_Iterative.setText(_translate("MainWindow", "Iterative"))
        self.checkBox.setText(_translate("MainWindow", "Max_Iter."))
        self.pushButton_Learn.setText(_translate("MainWindow", "Learn"))
        self.radioButton_WriteAllSamples.setText(_translate("MainWindow", "Write all samples"))
        # self.radioButton_DeepRBF.setText(_translate("MainWindow", "Deep RBF"))
        self.classificationGroupBox.setTitle(_translate("MainWindow", "Classification Option"))
        self.radioButton_Classify1.setText(_translate("MainWindow", "RBF - Best match"))
        self.radioButton_Classify2.setText(_translate("MainWindow", "RBF - Dominant"))
        self.radioButton_Classify3.setText(_translate("MainWindow", "RBF - Unanimity"))
        self.radioButton_Classify4.setText(_translate("MainWindow", "RBF - Min consensus of"))
        self.radioButton_Classify5.setText(_translate("MainWindow", "KNN - Best match"))
        self.radioButton_Classify6.setText(_translate("MainWindow", "KNN - Dominant"))
        self.radioButton_Classify7.setText(_translate("MainWindow", "KNN - Unanimity"))
        self.radioButton_Classify8.setText(_translate("MainWindow", "KNN - Min consensus of"))
        self.label_K.setText(_translate("MainWindow", "K"))
        self.pushButton_Classify.setText(_translate("MainWindow", "Classify"))
        # self.pushButton_Exit.setText(_translate("MainWindow", "Exit"))
        self.StatusBox.setTitle(_translate("MainWindow", "Status"))
        self.label_DName.setText(_translate("MainWindow", "Data Filename"))
        self.label_MnK.setText(_translate("MainWindow", "SimuNMnK"))
        self.label_n.setText(_translate("MainWindow", "n"))
        self.label_NNeurons.setText(_translate("MainWindow", "#_neurons"))
        self.pushButton_ForgetAll.setText(_translate("MainWindow", "Forget All"))
        self.label_UsedNeurons.setText(_translate("MainWindow", "#_Used neurons"))
        self.label_NSamples.setText(_translate("MainWindow", "#_samples"))
        self.ResultBox.setTitle(_translate("MainWindow", "Results"))
        self.pushButton_LoadDataset.setText(_translate("MainWindow", "Load \nLearn\ndataset"))
        self.pushButton_ExportDataset.setText(_translate("MainWindow", "Load \nclassify\ndataset"))
        self.pushButton_LoadKnowledge.setText(_translate("MainWindow", "Load \nknowledge"))
        self.pushButton_SaveKnowledge.setText(_translate("MainWindow", "Save \nKnowledge"))
        self.pushButton_Knowledge.setText(_translate("MainWindow", "Knowledge"))
        self.pushButton_DatasetInDetail.setText(_translate("MainWindow", "Dataset \nin detail"))
        self.pushButton_Exit.setText(_translate("MainWindow", "Exit"))
        self.pushButton_repeat.setText(_translate("MainWindow", "Repeat"))
        self.label_NSamples2.setText(_translate("MainWindow", "#_samples"))
        self.label_ID.setText(_translate("MainWindow", "ID"))
        self.label_UNCc.setText(_translate("MainWindow", "UNC_correct"))
        self.label_UNCi.setText(_translate("MainWindow", "UNC_incorrect"))
        self.label_UNK.setText(_translate("MainWindow", "UNK"))
        self.label_Total.setText(_translate("MainWindow", "Total"))
        self.label_MainGraph.setText(_translate("MainWindow", "Distribution per category"))
        self.label_NeuronsGraph.setText(_translate("MainWindow", "Distribution of neurons per category"))
        self.label_LearningCurveGraph.setText(_translate("MainWindow", "Learning curve"))
        self.label_AccuracyGraph.setText(_translate("MainWindow", "Accuracy per category"))
        self.label_PieGraph.setText(_translate("MainWindow", "Accuracy per Recognition Status"))
        self.label_autoCounter.setText(_translate("MainWindow", "Count :              / 100"))

     #이벤트 정의

    #Load dataset
    def showDialog(self):
        self.load_data = []
        fname=QFileDialog.getOpenFileName(self, 'Load dataset', "", "csv Files(*.csv)")
        #dataset이 선택되어졌을 경우 그 이후 사건들 실행할 수 있게 if 코드 추가
        if fname[0]:
            self.mainGraph.canvas.ax.clear()
            self.Browser_DName.clear()
            l=len(fname[0])
            for i in range (0,l):
                if "/" ==fname[0][i]:
                    temp =i
            temp+=1
            self.Browser_DName.append(fname[0][temp:])
            self.logBrowser.append(fname[0][temp:])
            f = open(fname[0], 'r')
            rdr = csv.reader(f)
            for line in rdr:
                if "PatternID" != line[0]:
                    data_temp = np.concatenate((line[2], line[3],line[5:]), axis=None)
                    self.load_data.append(data_temp)
            f.close()

            # Distribution per category graph
            # 입력 데이터 카테고리 카운팅
            in_val1 = 0
            in_val2 = 0
            in_val3 = 0
            in_val4 = 0
            self.data_num = len(self.load_data)
            num=str(self.data_num)
            self.Browser_NSamples.clear()
            self.Browser_NSamples.append(num)
            for i in range(0, self.data_num):
                if self.load_data[i][1] == '1':
                    in_val1 += 1
                elif self.load_data[i][1] == '2':
                    in_val2 += 1
                elif self.load_data[i][1] == '3':
                    in_val3 += 1
                elif self.load_data[i][1] == '4':
                    in_val4 += 1
            height_input = [in_val1, in_val2, in_val3, in_val4]
            bars = ('1', '2', '3', '4')
            y_pos = np.arange(len(bars)) + 0.8
            bar_width = 0.2
            # 입력 데이터 카테고리 갯수 막대 그래프 출력
            self.mainGraph.canvas.ax.bar(y_pos, height_input, width=bar_width, align='center', alpha=0.35, color='r', label='Input')
            self.mainGraph.canvas.ax.legend()
            self.mainGraph.canvas.draw()

        else:
            QMessageBox.about(self, "Warning", "No file selected.")

    def showDialog2(self):
        self.load_data2 = []
        fname=QFileDialog.getOpenFileName(self, 'Load dataset', "", "csv Files(*.csv)")
        #dataset이 선택되어졌을 경우 그 이후 사건들 실행할 수 있게 if 코드 추가
        if fname[0]:
            self.mainGraph.canvas.ax.clear()
            self.Browser_DName.clear()
            l=len(fname[0])
            for i in range (0,l):
                if "/" ==fname[0][i]:
                    temp =i
            temp+=1
            self.Browser_DName.append(fname[0][temp:])
            self.logBrowser.append(fname[0][temp:])
            f = open(fname[0], 'r')
            rdr = csv.reader(f)
            for line in rdr:
                if "PatternID" != line[0]:
                    data_temp = np.concatenate((line[2], line[3],line[5:]), axis=None)
                    self.load_data2.append(data_temp)
            f.close()

            # Distribution per category graph
            # 입력 데이터 카테고리 카운팅
            in_val1 = 0
            in_val2 = 0
            in_val3 = 0
            in_val4 = 0
            self.data_num2 = len(self.load_data2)
            num=str(self.data_num2)
            self.Browser_NSamples.clear()
            self.Browser_NSamples.append(num)
            for i in range(0, self.data_num2):
                if self.load_data2[i][1] == '1':
                    in_val1 += 1
                elif self.load_data2[i][1] == '2':
                    in_val2 += 1
                elif self.load_data2[i][1] == '3':
                    in_val3 += 1
                elif self.load_data2[i][1] == '4':
                    in_val4 += 1
            height_input = [in_val1, in_val2, in_val3, in_val4]
            bars = ('1', '2', '3', '4')
            y_pos = np.arange(len(bars)) + 0.8
            bar_width = 0.2
            # 입력 데이터 카테고리 갯수 막대 그래프 출력
            self.mainGraph.canvas.ax.bar(y_pos, height_input, width=bar_width, align='center', alpha=0.35, color='r', label='Input')
            self.mainGraph.canvas.ax.legend()
            self.mainGraph.canvas.draw()

        else:
            QMessageBox.about(self, "Warning", "No file selected.")

    def ForgetAll_clicked(self):
        self.neurons=0
        self.mainGraph.canvas.ax.clear()
        self.neuronsGraph.canvas.ax.clear()
        self.learningCurveGraph.canvas.ax.clear()
        self.accuracyCategoryBarGraph.canvas.ax.clear()
        self.accuracyPieGraph.canvas.ax.clear()
        self.mainGraph.canvas.draw()
        self.neuronsGraph.canvas.draw()
        self.learningCurveGraph.canvas.draw()
        self.accuracyCategoryBarGraph.canvas.draw()
        self.accuracyPieGraph.canvas.draw()

        self.Browser_DName.clear()
        self.logBrowser.clear()
        self.Browser_NSamples.clear()
        self.Browser_UsedNeurons.clear()
        self.Browser_NSamples2.clear()
        self.Browser_ID.clear()
        self.Browser_UNCc.clear()
        self.Browser_UNCi.clear()
        self.Browser_UNK.clear()
        self.Browser_Total.clear()

        self.capture()

    def AutoCounter(self):
        num=str(self.count)
        self.Browser_Count.setText(num)

    def spinBoxChanged(self):
        self.NMnKvalue = self.spinBox_SimuNMnK.value()
        self.maxifvalue = self.spinBox_maxif.value()
        self.minifvalue = self.spinBox_minif.value()
        self.itervalue = self.spinBox_Iter.value()
        self.MnKChanged()

    def MnKChanged(self):
        self.Browser_NNeurons.clear()
        value = str(self.NMnKvalue*1024)
        self.Browser_NNeurons.append(value)

    def rdobtn_clicked(self):
        # Iterative 선택
        self.rdo=0
        self.do=0
        if self.radioButton_Iterative.isChecked():
            self.rdo=1
            # Max_Iter. 선택시
            if self.checkBox.isChecked():
                self.chbox=1
            else:  # Max_Iter 선택안함
                self.chbox=0
        #Deep RBF 선택
        # elif self.radioButton_DeepRBF.isChecked():
        #     self.rdo=2
        # Write all samples 선택
        elif self.radioButton_WriteAllSamples.isChecked():
            self.rdo=9
        #Classify_RBF
        if self.radioButton_Classify1.isChecked():
            self.do = 3
        elif  self.radioButton_Classify2.isChecked() :
            self.do=3
        elif self.radioButton_Classify3.isChecked() :
            self.do=3
        elif self.radioButton_Classify4.isChecked():
            self.do=3
        #Classify_KNN
        elif self.radioButton_Classify5.isChecked():
            self.do = 4
        elif self.radioButton_Classify6.isChecked():
            self.do = 4
        elif self.radioButton_Classify7.isChecked():
            self.do = 4
        elif self.radioButton_Classify8.isChecked():
            self.do = 4

    def cat_out(self):
        rdo=0
        if self.radioButton_Classify1.isChecked():
            rdo=5
        elif self.radioButton_Classify2.isChecked():
            rdo=6
        elif self.radioButton_Classify3.isChecked():
            rdo=7
        elif self.radioButton_Classify4.isChecked():
            rdo=8
        elif self.radioButton_Classify5.isChecked():
            rdo=5
        elif self.radioButton_Classify6.isChecked():
            rdo=6
        elif self.radioButton_Classify7.isChecked():
            rdo=7
        elif self.radioButton_Classify8.isChecked():
            rdo=8
        return rdo

    def Learn_clicked(self):
        self.rdobtn_clicked()
        self.learnRBF()
        self.catCount()
        self.LearnSetBrowser()
    # ----------------------------------------

    def learnRBF(self):
        self.spinBoxChanged()
        self.neurons, self.iterval, self.ncount= test.learn_test(self.load_data, self.rdo, self.NMnKvalue, self.maxifvalue, self.minifvalue, num_board=1, verbose=1)
        self.logBrowser.append('Context 1 with Norm L1')
        radio=self.rdo
        val1=self.maxifvalue
        val3=self.minifvalue
        if radio==1:
            self.logBrowser.append('RBF(Iterative) 선택')
        elif radio==2:
            self.logBrowser.append('Write all samples 선택')
        self.logBrowser.append("MaxIF : {}, MinIF : {}".format(val1, val3))
        self.logBrowser.append("Iteration : {}".format(self.iterval))
        # print(self.ncount)
        print("self.neurons.neurons", self.neurons.neurons)

    # ----------------------------------------

    def LearnSetBrowser(self):
        self.Browser_UsedNeurons.clear()
        self.Browser_NSamples.clear()
        num=str(self.data_num)
        self.Browser_NSamples.append(num)
        temp=str(self.neurons.read_ncount())
        self.Browser_UsedNeurons.append(temp)

    def ClaSetBrowser(self):
        self.Browser_NSamples2.clear()
        self.Browser_ID.clear()
        self.Browser_UNCc.clear()
        self.Browser_UNCi.clear()
        self.Browser_UNK.clear()
        self.Browser_Total.clear()

        ID =str(self.ID)
        UNC_c = str(self.UNC_c)
        UNC_i = str(self.UNC_i)
        UKN = str(self.UKN)
        NSamples2=str(self.data_num2)

        w=self.ID*100/self.data_num2
        x=self.UNC_c * 100/ self.data_num2
        y=self.UNC_i*100 / self.data_num2
        z=self.UKN *100/ self.data_num2
        ID2=str(round(w,2))
        UNC_c2 = str(round(x,2) )
        UNC_i2 = str(round(y,2))
        UKN2 = str(round(z,2))
        self.Browser_NSamples2.append(NSamples2)
        self.Browser_ID.append("{} ({}%)".format(ID, ID2))
        self.Browser_UNCc.append("{} ({}%)".format(UNC_c, UNC_c2))
        self.Browser_UNCi.append("{} ({}%)".format(UNC_i, UNC_i2))
        self.Browser_UNK.append("{} ({}%)".format(UKN, UKN2))
        temp = (self.ID) + (self.UNC_c) + (self.UNC_i) + (self.UKN)
        total=str(temp)
        self.result=w+x
        self.Browser_Total.append(total)

    def catCount(self):
        # Distribution per category graph
        #Used_neurons 카테고리 카운팅
        data = self.neurons.neurons
        if data != 0 :
            cat_val1 = 0
            cat_val2 = 0
            cat_val3 = 0
            cat_val4 = 0
            i = 0
            while data[i].state == nrn.NeuronState.com :
                if data[i].cat == 1:
                    cat_val1 += 1
                elif data[i].cat == 2:
                    cat_val2 += 1
                elif data[i].cat == 3:
                    cat_val3 += 1
                elif data[i].cat == 4:
                    cat_val4 += 1
                i +=  1
            height_neuron = [cat_val1, cat_val2, cat_val3, cat_val4]
            xran=[1,2,3,4]
            line = self.ncount[0:(self.iterval+1)]
            label=[str(cat_val1), str(cat_val2), str(cat_val3), str(cat_val4)]
            bars = ('1', '2', '3', '4')
            y_po=np.arange(len(line))+1
            y_pos = np.arange(len(bars)) + 1
            bar_width=0.5
            self.mainGraph.canvas.ax.bar(y_pos, height_neuron, width=bar_width-0.3, align='center', alpha=0.35, color='gold', label='Used neurons')
            self.neuronsGraph.canvas.ax.clear()
            self.neuronsGraph.canvas.ax.bar(y_pos, height_neuron, width=bar_width, align='center', alpha=0.35, color='gold')
            self.learningCurveGraph.canvas.ax.stackplot(y_po, line, colors='lightsalmon')
            self.mainGraph.canvas.ax.legend()
            self.mainGraph.canvas.draw()
            self.learningCurveGraph.canvas.draw()
            for i in range(len(bars)):
                self.neuronsGraph.canvas.ax.text(x=xran[i], y=height_neuron[i], s=label[i], size=8)
            self.neuronsGraph.canvas.draw()
            # self.capture()
        else :
           print('to draw neuron graph is fail : data = 0')

    def SaveKnowledge(self):
        SK=openpyxl.Workbook()
        column_header=['NeuronID', 'Context', 'Category', 'InfluenceField', 'Degeneration', 'Distance', 'MinIF']
        sheet=SK.active
        sheet.append(column_header)
        data=self.neurons.neurons

        for i in range(0,1000):
            if data[i].state == nrn.NeuronState.com:
                id = str(data[i].id_)
                cxt = str(data[i].cxt)
                cat = str(data[i].cat)
                aif = str(data[i].aif)
                dist=str(data[i].dist)
                deg=data[i].degenerate
                sheet.cell(row=i+2, column=1).value=id
                sheet.cell(row=i+2, column=2).value = cxt
                sheet.cell(row=i+2, column=3).value = cat
                sheet.cell(row=i+2, column=4).value = aif
                sheet.cell(row=i + 2, column=5).value = deg
                sheet.cell(row=i + 2, column=6).value = dist
                sheet.cell(row=i+2, column=7).value = self.minifvalue
        SK.save('Knowledge2.xlsx')

    def LoadKnowledge(self):
        LK=openpyxl.load_workbook('Knowledge2.xlsx')
        sheet=LK.active
        val=1
        minval=2
        network = cm1k.CM1KEmulator(network_size=val* 1024)
        for i in range(0,11):
            temp = sheet.cell(row=i + 2, column=7).value
            if temp == minval:
                id_ = sheet.cell(row=i + 2, column=1).value
                cxt = sheet.cell(row=i + 2, column=2).value
                cat = sheet.cell(row=i + 2, column=3).value
                aif = sheet.cell(row=i + 2, column=4).value
                deg = sheet.cell(row=i + 2, column=4).value
                dist= sheet.cell(row=i + 2, column=6).value
                minif=sheet.cell(row=i + 2, column=7).value
                state = nrn.NeuronState.com
                cm1k.CM1KEmulator.read_nsr_mode=cm1k.CM1KMode.save_restore
                network.write_ncr_nid(id_)
                neuron.state = state
                network.write_ncr_context(cxt)
                network.write_cat(cat)
                network.write_aif(aif)
                network.write_cat_degenerate(deg)
                network.write_dist_non_ui(dist)
                network.write_minif(minif)



        self.Browser_UsedNeurons.clear()
        temp = str(self.neurons.read_ncount())
        self.Browser_UsedNeurons.append(temp)
        #그래프 리셋&그리기
        self.mainGraph.canvas.ax.clear()
        self.learningCurveGraph.canvas.ax.clear()
        self.neuronsGraph.canvas.ax.clear()
        self.mainGraph.canvas.draw()
        self.learningCurveGraph.canvas.draw()
        self.neuronsGraph.canvas.draw()
        self.catCount()


    #Used neurons content
    def viewKnowledge(self):
        self.setGeometry(200,300,380,400) #(x,y,width, height)
        self.knowledgeTable=QTableWidget(self)
        self.knowledgeTable.resize(370,390)
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("MainWindow", "Knowledge Content"))
        num = self.neurons.read_ncount()
        self.knowledgeTable.setRowCount(num)
        self.knowledgeTable.setColumnCount(5)
        self.setTableWidgetData()
        self.show()

    def setTableWidgetData(self):
        column_header=['NeuronID',  'Context',  'Category',  'InfluenceField',  'MinIF']
        self.knowledgeTable.setHorizontalHeaderLabels(column_header)
        data = self.neurons.neurons
        for i in range(0,1000):
            if data[i].state == nrn.NeuronState.com:
                #테이블 채우는 식
                id=str(data[i].id_)
                cxt=str(data[i].cxt)
                cat=str(data[i].cat)
                aif=str(data[i].aif)
                minif=str(self.minifvalue)
                item1=QTableWidgetItem(id)
                item2=QTableWidgetItem(cxt)
                item3=QTableWidgetItem(cat)
                item4=QTableWidgetItem(aif)
                item5=QTableWidgetItem(minif)
                item1.setTextAlignment(Qt.AlignCenter)
                item2.setTextAlignment(Qt.AlignCenter)
                item3.setTextAlignment(Qt.AlignCenter)
                item4.setTextAlignment(Qt.AlignCenter)
                item5.setTextAlignment(Qt.AlignCenter)
                self.knowledgeTable.setItem(i, 0, item1)
                self.knowledgeTable.setItem(i, 1, item2)
                self.knowledgeTable.setItem(i, 2, item3)
                self.knowledgeTable.setItem(i, 3, item4)
                self.knowledgeTable.setItem(i, 4, item5)
            else:
                break
        self.knowledgeTable.resizeColumnsToContents()
        #self.knowledge.resizeRowsToContents()

    def spin_value(self):
        Kvalue=self.spinBox_Kvalue.value()
        self.con_value=self.spinBox_MinConsensus.value()
        self.con_value2=self.spinBox_MinConsensus2.value()
        return Kvalue

    # ----------------------------------------

    def Classify_clicked(self):
        self.ID, self.UNC_c, self.UNC_i, self.UKN, self.total_detail, self.guess_cat, self.cat_accuracy = test.classify(self.load_data2, self.neurons, self.do, self.cat_out(), self.spin_value())
        self.getBar()
        self.pie()
        self.catAccuracyGraph()
        radio=self.do
        val1=self.maxifvalue
        val3=self.minifvalue
        catout=self.cat_out()
        val4 = self.spin_value()
        val5=self.con_value
        #radio button print in log browser
        if radio==3:
            if catout==5:
                self.logBrowser.append('Classify_RBF, rule=BestMatch')
            elif catout==6:
                self.logBrowser.append("Classify_RBF, rule=Dominant, K={}".format(val4))
            elif catout==7:
                self.logBrowser.append('Classify_RBF, rule=Unanimity')
            elif catout==8:
                self.logBrowser.append("Classify_RBF, rule=Min consensus of {}".format(val5))
        elif radio==4:
            self.logBrowser.append('Classify_KNN')
        self.ClaSetBrowser()
        self.count += 1
        self.AutoCounter()
        self.capture()
    # ----------------------------------------

    def getBar(self):
        #Distribution per category graph
        #테스트 후 판별한 카테고리 카운팅
        data=self.guess_cat
        te_val0 = 0
        te_val1 = 0
        te_val2 = 0
        te_val3 = 0
        te_val4 = 0
        data_num = len(data)
        for i in range(0, data_num):
            if data[i] == 1:
                te_val1 += 1
            elif data[i] == 2:
                te_val2 += 1
            elif data[i] == 3:
                te_val3 += 1
            elif data[i] == 4:
                te_val4 += 1
            elif data[i] ==None:
                te_val0 += 1
        height_output = [te_val0, te_val1, te_val2, te_val3, te_val4]
        bars = ('0', '1', '2', '3', '4')
        y_pos = np.arange(len(bars)) + 0.2
        bar_width = 0.2
        #테스트 데이터 카테고리 갯수 막대 그래프 출력
        self.mainGraph.canvas.ax.bar(y_pos, height_output, width=bar_width, align='center', alpha=0.35, color='mediumpurple', label='Output')
        self.mainGraph.canvas.ax.legend()
        self.mainGraph.canvas.draw()

    def catAccuracyGraph(self):
        cat_ = self.cat_accuracy
        idval1=0
        idval2=0
        idval3=0
        idval4=0
        unc_cval1=0
        unc_cval2 = 0
        unc_cval3 = 0
        unc_cval4 = 0
        unc_ival1=0
        unc_ival2 = 0
        unc_ival3 = 0
        unc_ival4 = 0
        unkval1=0
        unkval2 = 0
        unkval3 = 0
        unkval4 = 0
        # Accuracy per Category Graph
        for i in range(len(cat_)):
            if cat_[i][0] == 1:
                if cat_[i][1] == 'ID':
                    idval1+=1
                elif cat_[i][1] == 'UNC_c':
                    unc_cval1+=1
                elif cat_[i][1] == 'UNC_i':
                    unc_ival1+=1
                elif cat_[i][1] == 'UNK':
                    unkval1+=1
            elif cat_[i][0] == 2:
                if cat_[i][1] == 'ID':
                    idval2+=1
                elif cat_[i][1] == 'UNC_c':
                    unc_cval2+=1
                elif cat_[i][1] == 'UNC_i':
                    unc_ival2+=1
                elif cat_[i][1] == 'UNK':
                    unkval2+=1
            elif cat_[i][0] == 3:
                if cat_[i][1] == 'ID':
                    idval3 += 1
                elif cat_[i][1] == 'UNC_c':
                    unc_cval3 += 1
                elif cat_[i][1] == 'UNC_i':
                    unc_ival3 += 1
                elif cat_[i][1] == 'UNK':
                    unkval3 += 1
            elif cat_[i][0] == 4:
                if cat_[i][1] == 'ID':
                    idval4 += 1
                elif cat_[i][1] == 'UNC_c':
                    unc_cval4 += 1
                elif cat_[i][1] == 'UNC_i':
                    unc_ival4 += 1
                elif cat_[i][1] == 'UNK':
                    unkval4 += 1
        bars = ('1', '2', '3', '4')
        xran = [1, 2, 3, 4]
        id=[idval1, idval2, idval3, idval4]
        unc_c=[unc_cval1,unc_cval2,unc_cval3,unc_cval4]
        unc_i=[unc_ival1, unc_ival2, unc_ival3, unc_ival4]
        unk=[unkval1, unkval2, unkval3, unkval4]
        re1=id[0]+unc_c[0]+unc_i[0]+unk[0]
        re2 = id[1] + unc_c[1] + unc_i[1] + unk[1]
        re3 = id[2] + unc_c[2] + unc_i[2] + unk[2]
        re4 = id[3] + unc_c[3] + unc_i[3] + unk[3]
        number_label=[re1, re2, re3, re4]
        #str_label = [str(re1), str(re2), str(re3), str(re4)]
        y_pos = np.arange(len(bars))+1
        bar_width = 0.5
        #self.mainGraph.canvas.ax.bar(y_pos, number_label, width=bar_width, align='center', alpha=0.35, color='mediumpurple', label='Output')
        graph1=self.accuracyCategoryBarGraph.canvas.ax.bar(y_pos, id, width=bar_width, align='center', alpha=0.35, color='mediumseagreen')
        graph2=self.accuracyCategoryBarGraph.canvas.ax.bar(y_pos, unc_c, width=bar_width, align='center', alpha=0.35, color='dodgerblue')
        graph3=self.accuracyCategoryBarGraph.canvas.ax.bar(y_pos, unc_i, width=bar_width, align='center', alpha=0.35, color='gold')
        graph4=self.accuracyCategoryBarGraph.canvas.ax.bar(y_pos, unk, width=bar_width, align='center', alpha=0.35, color='crimson')
        self.accuracyCategoryBarGraph.canvas.ax.legend((graph1[0],graph2[0],graph3[0],graph4[0]),('ID', 'UNC_c', 'UNC_i', 'UNK'))
        self.accuracyCategoryBarGraph.canvas.draw()

    def pie(self):
        ID = self.ID
        UNC_c = self.UNC_c
        UNC_i = self.UNC_i
        UKN = self.UKN
        colors = ['yellowgreen', 'lightskyblue', 'gold', 'lightcoral']
        labels = ['ID', 'UNC_correct', 'UNC_incorrect', 'UKN']
        ratio = [ID, UNC_c, UNC_i, UKN]
        explode = (0.0, 0.1, 0.0, 0.0)
        self.accuracyPieGraph.canvas.ax.pie(ratio, explode=explode, colors=colors, shadow=True, startangle=90)
        self.accuracyPieGraph.canvas.draw()

    def viewDetail(self):
        self.setGeometry(200, 300, 700, 650)  # (x,y,width, height)
        self.detailAll = QTableWidget(self)
        self.detailAll.resize(690, 640)
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("MainWindow", "Detail Report(All)"))
        data_num = len(self.load_data2)
        self.detailAll.setRowCount(data_num)
        self.detailAll.setColumnCount(14)
        self.setTableWidgetData_Detail()
        self.show()

    def setTableWidgetData_Detail(self):
        column_header=['PatternID',  'Context',  'CatGT', 'Status', 'CatOut', 'Cat 1', 'Dist 1', 'Nid 1','Cat 2', 'Dist 2', 'Nid 2', 'Cat 3', 'Dist 3', 'Nid 3' ]
        self.detailAll.setHorizontalHeaderLabels(column_header)
        test_data = self.total_detail
        guesscat=self.guess_cat
        data_num = len(self.load_data2)

        for i in range(0,data_num):
            le = len(test_data[i][0])
            id = str(test_data[i][0][0])
            cxt = str(test_data[i][0][1])
            catgt = str(test_data[i][0][2])
            if le<=3:
                cat1 = None
                dist1 = None
                nid1 = None
                cat2 = None
                dist2 = None
                nid2 = None
                cat3 = None
                dist3 = None
                nid3 = None
            elif le>=4 and le<=6:
                cat1 = str(test_data[i][0][3])
                dist1 = str(test_data[i][0][4])
                nid1 = str(test_data[i][0][5])
                cat2 = None
                dist2 = None
                nid2 = None
                cat3 = None
                dist3 = None
                nid3 = None
            elif le>=7 and le<=9 :
                cat1 = str(test_data[i][0][3])
                dist1 = str(test_data[i][0][4])
                nid1 = str(test_data[i][0][5])
                cat2 = str(test_data[i][0][6])
                dist2 = str(test_data[i][0][7])
                nid2 = str(test_data[i][0][8])
                cat3 = None
                dist3 = None
                nid3 = None
            else:
                cat1 = str(test_data[i][0][3])
                dist1 = str(test_data[i][0][4])
                nid1 = str(test_data[i][0][5])
                cat2 = str(test_data[i][0][6])
                dist2 = str(test_data[i][0][7])
                nid2 = str(test_data[i][0][8])
                cat3 = str(test_data[i][0][9])
                dist3 = str(test_data[i][0][10])
                nid3 = str(test_data[i][0][11])
            catout = str(guesscat[i])
            if catgt==catout:
                sta = 'ID'
            elif catout == None:
                sta = 'Unknown'
            else:
                sta = 'Incorrect'
            item0 = QTableWidgetItem(id)
            item1 = QTableWidgetItem(cxt)
            item2 = QTableWidgetItem(catgt)
            item3 = QTableWidgetItem(sta)
            item4 = QTableWidgetItem(catout)
            item5 = QTableWidgetItem(cat1)
            item6 = QTableWidgetItem(dist1)
            item7 = QTableWidgetItem(nid1)
            item8 = QTableWidgetItem(cat2)
            item9 = QTableWidgetItem(dist2)
            item10 = QTableWidgetItem(nid2)
            item11 = QTableWidgetItem(cat3)
            item12 = QTableWidgetItem(dist3)
            item13 = QTableWidgetItem(nid3)
            item0.setTextAlignment(Qt.AlignCenter)
            item1.setTextAlignment(Qt.AlignCenter)
            item2.setTextAlignment(Qt.AlignCenter)
            item3.setTextAlignment(Qt.AlignCenter)
            item4.setTextAlignment(Qt.AlignCenter)
            item5.setTextAlignment(Qt.AlignCenter)
            item6.setTextAlignment(Qt.AlignCenter)
            item7.setTextAlignment(Qt.AlignCenter)
            item8.setTextAlignment(Qt.AlignCenter)
            item9.setTextAlignment(Qt.AlignCenter)
            item10.setTextAlignment(Qt.AlignCenter)
            item11.setTextAlignment(Qt.AlignCenter)
            item12.setTextAlignment(Qt.AlignCenter)
            item13.setTextAlignment(Qt.AlignCenter)
            self.detailAll.setItem(i, 0, item0)
            self.detailAll.setItem(i, 1, item1)
            self.detailAll.setItem(i, 2, item2)
            self.detailAll.setItem(i, 3, item3)
            self.detailAll.setItem(i, 4, item4)
            self.detailAll.setItem(i, 5, item5)
            self.detailAll.setItem(i, 6, item6)
            self.detailAll.setItem(i, 7, item7)
            self.detailAll.setItem(i, 8, item8)
            self.detailAll.setItem(i, 9, item9)
            self.detailAll.setItem(i, 10, item10)
            self.detailAll.setItem(i, 11, item11)
            self.detailAll.setItem(i, 12, item12)
            self.detailAll.setItem(i, 13, item13)

        self.detailAll.resizeColumnsToContents()


    def capture(self):
        img=ImageGrab.grab()
        if (self.i % 2)==0:
            name = int(self.i/2)
            saves="{}-1{}".format(name, '.png')
            img.save(saves)
        else:
            name = int(self.i/2)
            saves = "{}-2{}".format(name, '.png')
            img.save(saves)
        self.i += 1

    #시스템 종료
    def exitProgram(self):
        sys.exit(0)

#Matplotlib canvas class to create figure
class MplCanvas(Canvas):
    def __init__(self):
        self.fig=Figure()
        self.ax=self.fig.add_subplot(111)
        Canvas.__init__(self, self.fig)
        Canvas.setSizePolicy(self, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        Canvas.updateGeometry(self)
#Matplotlib widget
class MplWidget(QtWidgets.QWidget):
    def __init__(self, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.canvas=MplCanvas()
        self.vbl=QtWidgets.QVBoxLayout()
        self.vbl.addWidget(self.canvas)
        self.setLayout(self.vbl)

from pyqtgraph import PlotWidget
import sys
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
